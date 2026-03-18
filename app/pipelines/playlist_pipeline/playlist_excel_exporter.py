from __future__ import annotations

import re
import subprocess
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config.paths import FFPROBE_EXE


VIDEO_EXTS = {".mp4", ".mkv", ".mov", ".avi", ".webm"}


def format_duration(seconds: Any) -> str:
    """
    Convert seconds into:
    - M:SS
    - H:MM:SS (if duration is over 1 hour)
    """
    if seconds is None:
        return ""

    try:
        total = int(float(seconds))
    except Exception:
        return ""

    hours = total // 3600
    minutes = (total % 3600) // 60
    secs = total % 60

    if hours > 0:
        return f"{hours}:{minutes:02d}:{secs:02d}"

    return f"{minutes}:{secs:02d}"


def extract_video_id_from_filename(file_path: Path) -> str:
    """
    Extract YouTube video ID from filenames like:
    001 - Title [abc123XYZ].mp4
    """
    match = re.search(r"\[([^\]]+)\]\.[^.]+$", file_path.name)
    return match.group(1) if match else ""


def get_video_files(folder: Path) -> list[Path]:
    return sorted(
        [
            f for f in folder.glob("*")
            if f.is_file() and f.suffix.lower() in VIDEO_EXTS
        ]
    )


def probe_frame_size(video_path: Path) -> tuple[int | None, int | None]:
    """
    Uses ffprobe to get width and height of the first video stream.
    """
    try:
        cmd = [
            str(FFPROBE_EXE),
            "-v",
            "error",
            "-select_streams",
            "v:0",
            "-show_entries",
            "stream=width,height",
            "-of",
            "csv=p=0:s=x",
            str(video_path),
        ]

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=True,
        )

        value = result.stdout.strip()

        if not value or "x" not in value:
            return None, None

        width_str, height_str = value.split("x", 1)
        return int(width_str), int(height_str)

    except Exception:
        return None, None


def build_frame_size_label(width: int | None, height: int | None) -> str:
    """
    Returns something like:
    - 1080x1920 - Reel
    - 1920x1080 - Long Video
    """
    if not width or not height:
        return ""

    video_type = "Reel" if height > width else "Long Video"
    return f"{width}x{height} - {video_type}"


def read_transcript_text(transcript_path: Path) -> str:
    if not transcript_path.exists():
        return ""

    try:
        return transcript_path.read_text(encoding="utf-8").strip()
    except Exception:
        return ""


def autosize_worksheet(ws: Worksheet) -> None:
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)


def generate_playlist_excel(paths, manifest: dict[str, Any]) -> Path:
    """
    Generate an Excel file with these columns:
    - title
    - transcription
    - time
    - youtube_link
    - frame_size

    Output file:
    data/playlists/<slug>/metadata/<slug>_playlist_data.xlsx
    """
    entries = manifest.get("entries", []) or []
    videos_dir = Path(manifest["videos_dir"])
    transcripts_dir = Path(manifest["transcripts_dir"])

    video_files = get_video_files(videos_dir)

    # Map video_id -> downloaded video path
    video_map: dict[str, Path] = {}
    for video_file in video_files:
        video_id = extract_video_id_from_filename(video_file)
        if video_id:
            video_map[video_id] = video_file

    wb = Workbook()
    ws: Worksheet = cast(Worksheet, wb.active)
    ws.title = "Playlist Data"
    ws.freeze_panes = "A2"

    headers = [
        "title",
        "transcription",
        "time",
        "youtube_link",
        "frame_size",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for entry in entries:
        if not entry:
            continue

        video_id = entry.get("id", "")
        title = entry.get("title", "") or ""
        youtube_link = entry.get("url", "") or ""
        duration = format_duration(entry.get("duration"))

        video_file = video_map.get(video_id)
        transcript_text = ""
        frame_size = ""

        if video_file:
            transcript_file = transcripts_dir / f"{video_file.stem}.txt"
            transcript_text = read_transcript_text(transcript_file)

            width, height = probe_frame_size(video_file)
            frame_size = build_frame_size_label(width, height)

        ws.append([
            title,
            transcript_text,
            duration,
            youtube_link,
            frame_size,
        ])

    # Wrap text for transcript/title/link columns
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(vertical="top", wrap_text=True)  # title
        row[1].alignment = Alignment(vertical="top", wrap_text=True)  # transcription
        row[3].alignment = Alignment(vertical="top", wrap_text=True)  # youtube_link
        row[4].alignment = Alignment(vertical="top", wrap_text=True)  # frame_size

    autosize_worksheet(ws)

    # Make transcription column wider
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["A"].width = min(max(ws.column_dimensions["A"].width, 30), 50)
    ws.column_dimensions["D"].width = min(max(ws.column_dimensions["D"].width, 35), 60)
    ws.column_dimensions["E"].width = min(max(ws.column_dimensions["E"].width, 20), 30)

    output_file = paths.metadata / f"{paths.root.name}_playlist_data.xlsx"
    wb.save(output_file)

    return output_file