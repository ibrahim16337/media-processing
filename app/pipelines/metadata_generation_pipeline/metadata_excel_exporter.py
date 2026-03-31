from __future__ import annotations

from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


OUTPUT_HEADERS = ["filename", "title", "description", "tags", "hashtags"]


def _safe_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def autosize_worksheet(ws: Worksheet, max_width: int = 80) -> None:
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)


def export_metadata_excel(
    rows: list[dict[str, Any]],
    output_file: str | Path,
    sheet_name: str = "Metadata",
) -> Path:
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws: Worksheet = cast(Worksheet, wb.active)

    ws.title = sheet_name
    ws.freeze_panes = "A2"

    ws.append(OUTPUT_HEADERS)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in rows:
        ws.append(
            [
                _safe_string(row.get("filename", "")),
                _safe_string(row.get("title", "")),
                _safe_string(row.get("description", "")),
                _safe_string(row.get("tags", "")),
                _safe_string(row.get("hashtags", "")),
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    autosize_worksheet(ws, max_width=60)

    ws.column_dimensions["A"].width = min(max(float(ws.column_dimensions["A"].width or 0), 25), 40)
    ws.column_dimensions["B"].width = min(max(float(ws.column_dimensions["B"].width or 0), 35), 60)
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = min(max(float(ws.column_dimensions["D"].width or 0), 40), 80)
    ws.column_dimensions["E"].width = min(max(float(ws.column_dimensions["E"].width or 0), 25), 40)

    wb.save(output_path)
    return output_path