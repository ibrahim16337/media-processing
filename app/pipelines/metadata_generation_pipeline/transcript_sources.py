from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CANDIDATE_ENCODINGS = [
    "utf-8",
    "utf-8-sig",
    "utf-16",
    "utf-16-le",
    "utf-16-be",
    "cp1256",
    "iso-8859-6",
]

SUPPORTED_EXCEL_EXTS = {".xlsx", ".xlsm"}
SUPPORTED_TEXT_EXTS = {".txt"}


def read_text_any(path: Path) -> str:
    last_error: Exception | None = None

    for enc in CANDIDATE_ENCODINGS:
        try:
            return path.read_text(encoding=enc)
        except Exception as e:
            last_error = e

    if last_error:
        raise last_error

    raise RuntimeError(f"Failed to read file: {path}")


def _safe_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _build_transcript_item(
    filename: str,
    transcript: str,
    source_type: str,
    source_path: str,
    row_number: int | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    return {
        "filename": _safe_string(filename),
        "transcript": transcript or "",
        "source_type": source_type,
        "source_path": source_path,
        "row_number": row_number,
        "sheet_name": sheet_name or "",
    }


def load_single_transcript_file(file_path: str | Path) -> list[dict[str, Any]]:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Transcript file not found: {path}")

    if not path.is_file():
        raise ValueError(f"Expected a file path, got: {path}")

    if path.suffix.lower() not in SUPPORTED_TEXT_EXTS:
        raise ValueError(f"Unsupported transcript file type: {path.suffix}")

    transcript_text = read_text_any(path)

    return [
        _build_transcript_item(
            filename=path.name,
            transcript=transcript_text,
            source_type="single_file",
            source_path=str(path),
        )
    ]


def load_transcript_folder(folder_path: str | Path) -> list[dict[str, Any]]:
    folder = Path(folder_path)

    if not folder.exists():
        raise FileNotFoundError(f"Transcript folder not found: {folder}")

    if not folder.is_dir():
        raise ValueError(f"Expected a folder path, got: {folder}")

    files = sorted(
        [
            f for f in folder.glob("*.txt")
            if f.is_file() and f.suffix.lower() in SUPPORTED_TEXT_EXTS
        ]
    )

    items: list[dict[str, Any]] = []

    for file_path in files:
        transcript_text = read_text_any(file_path)
        items.append(
            _build_transcript_item(
                filename=file_path.name,
                transcript=transcript_text,
                source_type="folder",
                source_path=str(file_path),
            )
        )

    return items


def get_excel_sheet_names(excel_path: str | Path) -> list[str]:
    path = Path(excel_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    if path.suffix.lower() not in SUPPORTED_EXCEL_EXTS:
        raise ValueError(f"Unsupported Excel file type: {path.suffix}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def get_excel_columns(excel_path: str | Path, sheet_name: str | None = None) -> list[str]:
    path = Path(excel_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    if path.suffix.lower() not in SUPPORTED_EXCEL_EXTS:
        raise ValueError(f"Unsupported Excel file type: {path.suffix}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header_row = next(rows, None)

        if not header_row:
            return []

        return [_safe_string(cell) for cell in header_row]
    finally:
        wb.close()


def _column_letter_to_index(column_ref: str) -> int | None:
    column_ref = _safe_string(column_ref).upper()

    if not column_ref or not column_ref.isalpha():
        return None

    value = 0
    for char in column_ref:
        value = value * 26 + (ord(char) - ord("A") + 1)

    return value - 1


def _resolve_column_index(column_ref: str, headers: list[str]) -> int:
    column_ref = _safe_string(column_ref)

    if not column_ref:
        raise ValueError("Column reference cannot be empty.")

    letter_index = _column_letter_to_index(column_ref)
    if letter_index is not None:
        if 0 <= letter_index < len(headers):
            return letter_index
        raise ValueError(f"Excel column letter out of range: {column_ref}")

    normalized_headers = [h.strip().lower() for h in headers]
    target = column_ref.strip().lower()

    if target in normalized_headers:
        return normalized_headers.index(target)

    raise ValueError(f"Column not found in Excel headers: {column_ref}")


def load_transcripts_from_excel(
    excel_path: str | Path,
    transcript_column: str,
    filename_column: str | None = None,
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    path = Path(excel_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    if not path.is_file():
        raise ValueError(f"Expected a file path, got: {path}")

    if path.suffix.lower() not in SUPPORTED_EXCEL_EXTS:
        raise ValueError(f"Unsupported Excel file type: {path.suffix}")

    wb = load_workbook(path, read_only=True, data_only=True)

    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)

        if not header_row:
            return []

        headers = [_safe_string(cell) for cell in header_row]

        transcript_idx = _resolve_column_index(transcript_column, headers)
        filename_idx = None
        if filename_column:
            filename_idx = _resolve_column_index(filename_column, headers)

        items: list[dict[str, Any]] = []
        item_counter = 0

        for excel_row_number, row in enumerate(row_iter, start=2):
            row_values = list(row)

            transcript_text = ""
            if transcript_idx < len(row_values):
                transcript_text = _safe_string(row_values[transcript_idx])

            if not transcript_text:
                continue

            item_counter += 1

            if filename_idx is not None and filename_idx < len(row_values):
                filename_value = _safe_string(row_values[filename_idx])
            else:
                filename_value = ""

            if not filename_value:
                filename_value = f"row_{item_counter:03d}"

            items.append(
                _build_transcript_item(
                    filename=filename_value,
                    transcript=transcript_text,
                    source_type="excel",
                    source_path=str(path),
                    row_number=excel_row_number,
                    sheet_name=ws.title,
                )
            )

        return items

    finally:
        wb.close()