from __future__ import annotations

import time
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

from app.config.paths import (
    METADATA_BATCH_DIR,
    METADATA_EXCEL_IMPORT_DIR,
    METADATA_PLAYLIST_DIR,
    METADATA_SINGLE_DIR,
    PLAYLISTS_DIR,
    TRANSCRIPT_DIR,
    UPLOAD_AUDIO_DIR,
    UPLOAD_VIDEO_DIR,
    slugify,
)
from app.pipelines.export_pipeline.metadata_excel_exporter import (
    export_metadata_excel,
)
from app.pipelines.export_pipeline.metadata_json_exporter import (
    export_metadata_json,
)
from app.pipelines.metadata_generation_pipeline.ollama_client import (
    DEFAULT_BASE_URL,
    DEFAULT_MODEL,
    DEFAULT_NUM_CTX,
    DEFAULT_NUM_PREDICT,
    DEFAULT_RETRIES,
    DEFAULT_SLEEP_MS,
    DEFAULT_TEMPERATURE,
    DEFAULT_TIMEOUT,
    generate_metadata_from_prompt,
)
from app.pipelines.metadata_generation_pipeline.prompt_builder import (
    build_metadata_prompt,
)
from app.pipelines.metadata_generation_pipeline.response_parser import (
    parse_metadata_json,
)
from app.pipelines.metadata_generation_pipeline.transcript_sources import (
    load_single_transcript_file,
    load_transcript_folder,
    load_transcripts_from_excel,
)
from app.pipelines.workflow_pipeline.transcript_export_workflows import (
    build_transcript_rows,
)


SUPPORTED_SOURCE_TYPES = {"single_file", "folder", "excel"}


def _safe_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_transcript_text(text: str) -> str:
    text = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")

    cleaned_lines: list[str] = []
    header_phase = True

    for line in lines:
        stripped = line.strip()

        if header_phase:
            if not stripped:
                continue
            if stripped.startswith("Audio Length:"):
                continue
            if stripped.startswith("Detected Language:"):
                continue

            header_phase = False

        cleaned_lines.append(line)

    return "\n".join(cleaned_lines).strip()


def _timestamp_str() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text or "", encoding="utf-8")


def _append_tsv_row(path: Path, values: list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    line = "\t".join(_safe_string(v).replace("\t", " ").replace("\n", " ") for v in values) + "\n"
    with path.open("a", encoding="utf-8") as f:
        f.write(line)


def _path_is_under(child: Path, parent: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except Exception:
        return False


def _resolve_base_output_dir(source_type: str, source_path: str | Path) -> Path:
    source_path_obj = Path(source_path)

    if _path_is_under(source_path_obj, PLAYLISTS_DIR):
        return METADATA_PLAYLIST_DIR

    if source_type == "single_file":
        return METADATA_SINGLE_DIR
    if source_type == "folder":
        return METADATA_BATCH_DIR
    if source_type == "excel":
        return METADATA_EXCEL_IMPORT_DIR

    raise ValueError(f"Unsupported source type: {source_type}")


def _load_source_items(
    source_type: str,
    source_path: str | Path,
    transcript_column: str | None = None,
    filename_column: str | None = None,
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    if source_type == "single_file":
        return load_single_transcript_file(source_path)

    if source_type == "folder":
        return load_transcript_folder(source_path)

    if source_type == "excel":
        if not transcript_column:
            raise ValueError("transcript_column is required when source_type='excel'.")

        return load_transcripts_from_excel(
            excel_path=source_path,
            transcript_column=transcript_column,
            filename_column=filename_column,
            sheet_name=sheet_name,
        )

    raise ValueError(f"Unsupported source type: {source_type}")


def _build_run_dir(
    source_type: str,
    source_path: str | Path,
    output_name: str | None = None,
    output_dir: str | Path | None = None,
) -> Path:
    if output_dir:
        base_dir = Path(output_dir)
    else:
        base_dir = _resolve_base_output_dir(source_type, source_path)

    source_path_obj = Path(source_path)
    default_name = source_path_obj.stem if source_path_obj.is_file() else source_path_obj.name
    run_name = slugify(output_name or default_name or source_type)
    run_dir = base_dir / f"{_timestamp_str()}_{run_name}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def _resolve_media_lookup_dirs_for_transcript_source(
    source_type: str,
    source_path: str | Path,
) -> list[Path]:
    source_path_obj = Path(source_path)
    lookup_dirs: list[Path] = []

    if source_type == "single_file":
        parent = source_path_obj.parent

        if _path_is_under(source_path_obj, TRANSCRIPT_DIR):
            lookup_dirs.extend([UPLOAD_AUDIO_DIR, UPLOAD_VIDEO_DIR])

        if parent.name.lower() == "transcripts":
            playlist_root = parent.parent
            lookup_dirs.extend([playlist_root / "audio", playlist_root / "videos"])

    elif source_type == "folder":
        folder = source_path_obj

        if _path_is_under(folder, TRANSCRIPT_DIR):
            lookup_dirs.extend([UPLOAD_AUDIO_DIR, UPLOAD_VIDEO_DIR])

        if folder.name.lower() == "transcripts":
            playlist_root = folder.parent
            lookup_dirs.extend([playlist_root / "audio", playlist_root / "videos"])

    unique_dirs: list[Path] = []
    seen: set[str] = set()

    for path in lookup_dirs:
        path_key = str(path.resolve()) if path.exists() else str(path)
        if path_key not in seen:
            seen.add(path_key)
            unique_dirs.append(path)

    return unique_dirs


def _build_transcript_enrichment_by_filename(
    source_type: str,
    source_path: str | Path,
) -> dict[str, dict[str, str]]:
    if source_type == "single_file":
        transcript_files = [Path(source_path)]
    elif source_type == "folder":
        transcript_files = sorted(Path(source_path).glob("*.txt"))
    else:
        return {}

    media_lookup_dirs = _resolve_media_lookup_dirs_for_transcript_source(source_type, source_path)

    rows = build_transcript_rows(
        transcript_files=transcript_files,
        media_lookup_dirs=media_lookup_dirs,
    )

    return {row["filename"]: row for row in rows}


def _parse_video_type_from_frame_size(frame_size_value: str) -> str:
    value = _safe_string(frame_size_value).lower()

    if not value:
        return ""

    if "reel" in value or "portrait" in value or "short" in value:
        return "Short"

    if "long" in value or "horizontal" in value:
        return "Long"

    return ""


def _find_header_index(headers: Sequence[str], target_name: str) -> int | None:
    normalized_headers = [h.strip().lower() for h in headers]
    target = target_name.strip().lower()

    if target in normalized_headers:
        return normalized_headers.index(target)

    return None


def _build_excel_enrichment_by_row_number(
    excel_path: str | Path,
    transcript_column: str,
    sheet_name: str | None = None,
) -> dict[int, dict[str, str]]:
    path = Path(excel_path)

    if not path.exists() or not path.is_file():
        return {}

    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)

        if not header_row:
            return {}

        headers = [_safe_string(cell) for cell in header_row]
        transcript_idx = _find_header_index(headers, transcript_column)
        time_idx = _find_header_index(headers, "time")
        frame_size_idx = _find_header_index(headers, "frame_size")

        enrichment: dict[int, dict[str, str]] = {}

        for excel_row_number, row in enumerate(row_iter, start=2):
            row_values = list(row)

            transcript_text = ""
            if transcript_idx is not None and transcript_idx < len(row_values):
                transcript_text = _clean_transcript_text(_safe_string(row_values[transcript_idx]))

            video_length = ""
            if time_idx is not None and time_idx < len(row_values):
                video_length = _safe_string(row_values[time_idx])

            video_type = ""
            if frame_size_idx is not None and frame_size_idx < len(row_values):
                video_type = _parse_video_type_from_frame_size(_safe_string(row_values[frame_size_idx]))

            enrichment[excel_row_number] = {
                "transcript": transcript_text,
                "video_length": video_length,
                "video_type": video_type,
            }

        return enrichment

    finally:
        workbook.close()


def _build_enriched_metadata_row(
    item: dict[str, Any],
    parsed_data: dict[str, str],
    transcript_enrichment_by_filename: dict[str, dict[str, str]] | None = None,
    excel_enrichment_by_row_number: dict[int, dict[str, str]] | None = None,
) -> dict[str, str]:
    filename = _safe_string(item.get("filename", ""))
    transcript_text = _clean_transcript_text(_safe_string(item.get("transcript", "")))
    video_length = ""
    video_type = ""

    if transcript_enrichment_by_filename:
        enrichment = transcript_enrichment_by_filename.get(filename, {})
        transcript_text = _clean_transcript_text(_safe_string(enrichment.get("transcription", transcript_text)))
        video_length = _safe_string(enrichment.get("audio length", ""))
        video_type = _safe_string(enrichment.get("video type", ""))

    if excel_enrichment_by_row_number:
        row_number = item.get("row_number")
        if isinstance(row_number, int):
            enrichment = excel_enrichment_by_row_number.get(row_number, {})
            transcript_text = _clean_transcript_text(_safe_string(enrichment.get("transcript", transcript_text)))
            video_length = _safe_string(enrichment.get("video_length", video_length))
            video_type = _safe_string(enrichment.get("video_type", video_type))

    return {
        "filename": filename,
        "title": _safe_string(parsed_data.get("title", "")),
        "transcript": transcript_text,
        "video_length": video_length,
        "video_type": video_type,
        "description": _safe_string(parsed_data.get("description", "")),
        "tags": _safe_string(parsed_data.get("tags", "")),
        "hashtags": _safe_string(parsed_data.get("hashtags", "")),
        "upload_link": "",
    }


def run_metadata_generation(
    source_type: str,
    source_path: str | Path,
    transcript_column: str | None = None,
    filename_column: str | None = None,
    sheet_name: str | None = None,
    output_name: str | None = None,
    output_dir: str | Path | None = None,
    model: str = DEFAULT_MODEL,
    base_url: str = DEFAULT_BASE_URL,
    timeout: int = DEFAULT_TIMEOUT,
    retries: int = DEFAULT_RETRIES,
    sleep_ms: int = DEFAULT_SLEEP_MS,
    temperature: float = DEFAULT_TEMPERATURE,
    num_ctx: int = DEFAULT_NUM_CTX,
    num_predict: int = DEFAULT_NUM_PREDICT,
    seed: int | None = None,
    system_message: str = "ہمیشہ ان پٹ کی زبان کا احترام کریں اور اگر ان پٹ اردو میں ہو تو اردو میں واضح، درست اور باوقار انداز میں جواب دیں۔",
) -> dict[str, Any]:
    source_type = _safe_string(source_type)

    if source_type not in SUPPORTED_SOURCE_TYPES:
        raise ValueError(
            f"Unsupported source_type: {source_type}. "
            f"Expected one of: {', '.join(sorted(SUPPORTED_SOURCE_TYPES))}"
        )

    items = _load_source_items(
        source_type=source_type,
        source_path=source_path,
        transcript_column=transcript_column,
        filename_column=filename_column,
        sheet_name=sheet_name,
    )

    run_dir = _build_run_dir(
        source_type=source_type,
        source_path=source_path,
        output_name=output_name,
        output_dir=output_dir,
    )

    raw_dir = run_dir / "raw_responses"
    raw_dir.mkdir(parents=True, exist_ok=True)

    ok_log = run_dir / "llm_ok.tsv"
    err_log = run_dir / "llm_err.tsv"
    excel_output = run_dir / "metadata_output.xlsx"
    json_output = run_dir / "metadata_output.json"

    ok_log.write_text("status\tindex\tfilename\tchars\tattempt\n", encoding="utf-8")
    err_log.write_text("status\tindex\tfilename\terror\n", encoding="utf-8")

    transcript_enrichment_by_filename: dict[str, dict[str, str]] = {}
    excel_enrichment_by_row_number: dict[int, dict[str, str]] = {}

    if source_type in {"single_file", "folder"}:
        transcript_enrichment_by_filename = _build_transcript_enrichment_by_filename(
            source_type=source_type,
            source_path=source_path,
        )

    if source_type == "excel" and transcript_column:
        excel_enrichment_by_row_number = _build_excel_enrichment_by_row_number(
            excel_path=source_path,
            transcript_column=transcript_column,
            sheet_name=sheet_name,
        )

    rows: list[dict[str, str]] = []
    errors: list[dict[str, Any]] = []

    if not items:
        export_metadata_excel(rows=[], output_file=excel_output)
        export_metadata_json(rows=[], output_file=json_output)

        return {
            "ok": True,
            "source_type": source_type,
            "source_path": str(source_path),
            "run_dir": str(run_dir),
            "excel_output": str(excel_output),
            "json_output": str(json_output),
            "ok_log": str(ok_log),
            "err_log": str(err_log),
            "total_items": 0,
            "success_count": 0,
            "error_count": 0,
            "rows": [],
            "errors": [],
        }

    total_items = len(items)

    for index, item in enumerate(items, start=1):
        filename = _safe_string(item.get("filename", f"item_{index:03d}"))
        transcript_text = _clean_transcript_text(_safe_string(item.get("transcript", "")))

        prompt = build_metadata_prompt(transcript_text)

        llm_result = generate_metadata_from_prompt(
            user_prompt=prompt,
            model=model,
            base_url=base_url,
            timeout=timeout,
            retries=retries,
            sleep_ms=sleep_ms,
            temperature=temperature,
            num_ctx=num_ctx,
            num_predict=num_predict,
            seed=seed,
            system_message=system_message,
        )

        raw_response_text = _safe_string(llm_result.get("content", ""))
        raw_response_file = raw_dir / f"{index:03d}_{slugify(Path(filename).stem or filename)}_response.txt"
        _write_text(raw_response_file, raw_response_text)

        if not llm_result.get("ok", False):
            error_message = _safe_string(llm_result.get("error", "Unknown Ollama error"))
            _append_tsv_row(err_log, ["HTTPERR", index, filename, error_message])

            rows.append(
                _build_enriched_metadata_row(
                    item=item,
                    parsed_data={
                        "title": "",
                        "description": "",
                        "tags": "",
                        "hashtags": "",
                    },
                    transcript_enrichment_by_filename=transcript_enrichment_by_filename,
                    excel_enrichment_by_row_number=excel_enrichment_by_row_number,
                )
            )

            errors.append(
                {
                    "index": index,
                    "filename": filename,
                    "stage": "ollama_call",
                    "error": error_message,
                    "raw_response_file": str(raw_response_file),
                }
            )

            if sleep_ms > 0 and index < total_items:
                time.sleep(sleep_ms / 1000.0)
            continue

        parsed = parse_metadata_json(raw_response_text)
        parsed_data = parsed.get("data", {}) or {}

        rows.append(
            _build_enriched_metadata_row(
                item=item,
                parsed_data=parsed_data,
                transcript_enrichment_by_filename=transcript_enrichment_by_filename,
                excel_enrichment_by_row_number=excel_enrichment_by_row_number,
            )
        )

        if parsed.get("ok", False):
            _append_tsv_row(
                ok_log,
                [
                    "OK",
                    index,
                    filename,
                    len(raw_response_text),
                    llm_result.get("attempt", ""),
                ],
            )
        else:
            parse_error = _safe_string(parsed.get("error", "Invalid metadata JSON"))
            _append_tsv_row(err_log, ["PARSEERR", index, filename, parse_error])

            errors.append(
                {
                    "index": index,
                    "filename": filename,
                    "stage": "response_parse",
                    "error": parse_error,
                    "raw_response_file": str(raw_response_file),
                }
            )

        if sleep_ms > 0 and index < total_items:
            time.sleep(sleep_ms / 1000.0)

    export_metadata_excel(rows=rows, output_file=excel_output)
    export_metadata_json(rows=rows, output_file=json_output)

    return {
        "ok": True,
        "source_type": source_type,
        "source_path": str(source_path),
        "run_dir": str(run_dir),
        "excel_output": str(excel_output),
        "json_output": str(json_output),
        "ok_log": str(ok_log),
        "err_log": str(err_log),
        "raw_dir": str(raw_dir),
        "total_items": total_items,
        "success_count": total_items - len(errors),
        "error_count": len(errors),
        "rows": rows,
        "errors": errors,
        "config": {
            "model": model,
            "base_url": base_url,
            "timeout": timeout,
            "retries": retries,
            "sleep_ms": sleep_ms,
            "temperature": temperature,
            "num_ctx": num_ctx,
            "num_predict": num_predict,
            "seed": seed,
        },
    }